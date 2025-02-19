# 导入库
import hashlib
import json
import random
import time
import uuid

import requests
import tls_client
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

GLOBALS = {
    "AD": "安道尔",
    "AE": "阿联酋",
    "AF": "阿富汗",
    "AG": "安提瓜和巴布达",
    "AI": "安圭拉",
    "AL": "阿尔巴尼亚",
    "AM": "亚美尼亚",
    "AO": "安哥拉",
    "AR": "阿根廷",
    "AS": "美属萨摩亚",
    "AT": "奥地利",
    "AU": "澳大利亚",
    "AW": "阿鲁巴",
    "AX": "奥兰群岛",
    "AZ": "阿塞拜疆",
    "BA": "波斯尼亚和黑塞哥维那",
    "BB": "巴巴多斯",
    "BD": "孟加拉国",
    "BE": "比利时",
    "BF": "布基纳法索",
    "BG": "保加利亚",
    "BH": "巴林",
    "BI": "布隆迪",
    "BJ": "贝尼恩",
    "BL": "圣巴泰勒米岛",
    "BM": "百慕大",
    "BN": "文莱",
    "BO": "玻利维亚",
    "BQ": "加勒比荷兰",
    "BR": "巴西",
    "BS": "巴哈马",
    "BT": "不丹",
    "BV": "布韦岛",
    "BW": "博茨瓦纳",
    "BY": "白俄罗斯",
    "BZ": "伯利兹",
    "CA": "加拿大",
    "CC": "科科斯（基林）群岛",
    "CD": "刚果（金）",
    "CF": "中非",
    "CG": "刚果（布）",
    "CH": "瑞士",
    "CI": "科特迪瓦",
    "CK": "库克群岛",
    "CL": "智利",
    "CM": "喀麦隆",
    "CN": "中国",
    "CO": "哥伦比亚",
    "CR": "哥斯达黎加",
    "CU": "古巴",
    "CV": "佛得角",
    "CW": "库拉索",
    "CX": "圣诞岛",
    "CY": "塞浦路斯",
    "CZ": "捷克",
    "DE": "德国",
    "DJ": "吉布提",
    "DK": "丹麦",
    "DM": "多米尼克",
    "DO": "多米尼加共和国",
    "DZ": "阿尔及利亚",
    "EC": "厄瓜多尔",
    "EE": "爱沙尼亚",
    "EG": "埃及",
    "EH": "西撒哈拉",
    "ER": "厄立特里亚",
    "ES": "西班牙",
    "ET": "埃塞俄比亚",
    "FI": "芬兰",
    "FJ": "斐济",
    "FM": "密克罗尼西亚联邦",
    "FO": "法罗群岛",
    "FR": "法国",
    "GA": "加蓬",
    "GB": "英国",
    "GD": "格林纳达",
    "GE": "格鲁吉亚",
    "GF": "法属圭亚那",
    "GG": "根西岛",
    "GH": "加纳",
    "GI": "直布罗陀",
    "GL": "格陵兰",
    "GM": "冈比亚",
    "GN": "几内亚",
    "GP": "瓜德罗普岛",
    "GQ": "赤道几内亚",
    "GR": "希腊",
    "GT": "危地马拉",
    "GU": "关岛",
    "GW": "几内亚比绍",
    "GY": "圭亚那",
    "HK": "香港",
    "HM": "赫德岛和麦克唐纳群岛",
    "HN": "洪都拉斯",
    "HR": "克罗地亚",
    "HT": "海地",
    "HU": "匈牙利",
    "ID": "印度尼西亚",
    "IE": "爱尔兰",
    "IL": "以色列",
    "IM": "曼岛",
    "IN": "印度",
    "IO": "英属印度洋领地",
    "IQ": "伊拉克",
    "IR": "伊朗",
    "IS": "冰岛",
    "IT": "意大利",
    "JE": "泽西岛",
    "JM": "牙买加",
    "JO": "约旦",
    "JP": "日本",
    "KE": "肯尼亚",
    "KG": "吉尔吉斯斯坦",
    "KH": "柬埔寨",
    "KI": "基里巴斯",
    "KM": "科摩罗",
    "KN": "圣基茨和尼维斯",
    "KP": "朝鲜",
    "KR": "韩国",
    "KW": "科威特",
    "KY": "开曼群岛",
    "KZ": "哈萨克斯坦",
    "LA": "老挝",
    "LB": "黎巴嫩",
    "LC": "圣卢西亚",
    "LI": "列支敦士登",
    "LK": "斯里兰卡",
    "LR": "利比里亚",
    "LS": "莱索托",
    "LT": "立陶宛",
    "LU": "卢森堡",
    "LV": "拉脱维亚",
    "LY": "利比亚",
    "MA": "摩洛哥",
    "MC": "摩纳哥",
    "MD": "摩尔多瓦",
    "ME": "黑山",
    "MF": "法属圣马丁岛",
    "MG": "马达加斯加",
    "MH": "马绍尔群岛",
    "MK": "北马其顿",
    "ML": "马里",
    "MM": "缅甸",
    "MN": "蒙古",
    "MO": "澳门",
    "MP": "北马里亚纳群岛",
    "MQ": "马提尼克岛",
    "MR": "毛里塔尼亚",
    "MS": "蒙特塞拉特",
    "MT": "马耳他",
    "MU": "毛里求斯",
    "MV": "马尔代夫",
    "MW": "马拉维",
    "MX": "墨西哥",
    "MY": "马来西亚",
    "MZ": "莫桑比克",
    "NA": "纳米比亚",
    "NC": "新喀里多尼亚",
    "NE": "尼日尔",
    "NF": "诺福克岛",
    "NG": "尼日利亚",
    "NI": "尼加拉瓜",
    "NL": "荷兰",
    "NO": "挪威",
    "NP": "尼泊尔",
    "NR": "瑙鲁",
    "NU": "纽埃",
    "NZ": "新西兰",
    "OM": "阿曼",
    "PA": "巴拿马",
    "PE": "秘鲁",
    "PF": "法属波利尼西亚",
    "PG": "巴布亚新几内亚",
    "PH": "菲律宾",
    "PK": "巴基斯坦",
    "PL": "波兰",
    "PM": "圣皮埃尔和密克隆",
    "PN": "皮特凯恩群岛",
    "PR": "波多黎各",
    "PT": "葡萄牙",
    "PW": "帕劳",
    "PY": "巴拉圭",
    "QA": "卡塔尔",
    "RE": "留尼汪",
    "RO": "罗马尼亚",
    "RS": "塞尔维亚",
    "RU": "俄罗斯",
    "RW": "卢旺达",
    "SA": "沙特阿拉伯",
    "SB": "所罗门群岛",
    "SC": "塞舌尔",
    "SD": "苏丹",
    "SE": "瑞典",
    "SG": "新加坡",
    "SH": "圣赫勒拿岛",
    "SI": "斯洛文尼亚",
    "SJ": "斯瓦尔巴群岛和扬马延岛",
    "SK": "斯洛伐克",
    "SL": "塞拉利昂",
    "SM": "圣马力诺",
    "SN": "塞内加尔",
    "SO": "索马里",
    "SR": "苏里南",
    "SS": "南苏丹",
    "ST": "圣多美和普林西比",
    "SV": "萨尔瓦多",
    "SX": "荷属圣马丁",
    "SY": "叙利亚",
    "SZ": "斯威士兰",
    "TC": "特克斯和凯科斯群岛",
    "TD": "乍得",
    "TF": "法属南部领土",
    "TG": "多哥",
    "TH": "泰国",
    "TJ": "塔吉克斯坦",
    "TK": "托克劳",
    "TL": "东帝汶",
    "TM": "土库曼斯坦",
    "TN": "突尼斯",
    "TO": "汤加",
    "TR": "土耳其",
    "TT": "特立尼达和多巴哥",
    "TV": "图瓦卢",
    "TZ": "坦桑尼亚",
    "UA": "乌克兰",
    "UG": "乌干达",
    "US": "美国",
    "UY": "乌拉圭",
    "UZ": "乌兹别克斯坦",
    "VA": "梵蒂冈",
    "VC": "圣文森特和格林纳丁斯",
    "VE": "委内瑞拉",
    "VG": "英属维尔京群岛",
    "VI": "美属维尔京群岛",
    "VN": "越南",
    "VU": "瓦努阿图",
    "WF": "瓦利斯和富图纳群岛",
    "WS": "萨摩亚",
    "YE": "也门",
    "YT": "马约特",
    "ZA": "南非",
    "ZM": "赞比亚",
    "ZW": "津巴布韦"
}
ASIA = {
    "AE": "阿联酋",
    "AF": "阿富汗",
    "AM": "亚美尼亚",
    "AZ": "阿塞拜疆",
    "BH": "巴林",
    "BD": "孟加拉国",
    "BT": "不丹",
    "BN": "文莱",
    "KH": "柬埔寨",
    "CN": "中国",
    "CY": "塞浦路斯",
    "GE": "格鲁吉亚",
    "IN": "印度",
    "ID": "印度尼西亚",
    "IR": "伊朗",
    "IQ": "伊拉克",
    "IL": "以色列",
    "JO": "约旦",
    "JP": "日本",
    "KG": "吉尔吉斯斯坦",
    "KP": "朝鲜",
    "KR": "韩国",
    "KW": "科威特",
    "KZ": "哈萨克斯坦",
    "LA": "老挝",
    "LB": "黎巴嫩",
    "MY": "马来西亚",
    "MV": "马尔代夫",
    "MN": "蒙古",
    "MM": "缅甸",
    "NP": "尼泊尔",
    "OM": "阿曼",
    "PK": "巴基斯坦",
    "PH": "菲律宾",
    "QA": "卡塔尔",
    "RU": "俄罗斯",
    "SA": "沙特阿拉伯",
    "SG": "新加坡",
    "LK": "斯里兰卡",
    "SY": "叙利亚",
    "TJ": "塔吉克斯坦",
    "TH": "泰国",
    "TR": "土耳其",
    "TM": "土库曼斯坦",
    "UZ": "乌兹别克斯坦",
    "VN": "越南",
    "YE": "也门"
}
OCEANIA = {
    "AU": "澳大利亚",
    "FJ": "斐济",
    "FM": "密克罗尼西亚联邦",
    "GU": "关岛",
    "KI": "基里巴斯",
    "MH": "马绍尔群岛",
    "MP": "北马里亚纳群岛",
    "NC": "新喀里多尼亚",
    "NZ": "新西兰",
    "PW": "帕劳",
    "PG": "巴布亚新几内亚",
    "WS": "萨摩亚",
    "TV": "图瓦卢",
    "VU": "瓦努阿图",
    "WF": "瓦利斯和富图纳群岛"
}
AFRICA = {
    "DZ": "阿尔及利亚",
    "AO": "安哥拉",
    "BJ": "贝尼恩",
    "BW": "博茨瓦纳",
    "BF": "布基纳法索",
    "BI": "布隆迪",
    "CM": "喀麦隆",
    "CV": "佛得角",
    "CF": "中非",
    "TD": "乍得",
    "KM": "科摩罗",
    "CG": "刚果（布）",
    "CD": "刚果（金）",
    "DJ": "吉布提",
    "EG": "埃及",
    "GQ": "赤道几内亚",
    "ER": "厄立特里亚",
    "ET": "埃塞俄比亚",
    "GA": "加蓬",
    "GM": "冈比亚",
    "GH": "加纳",
    "GN": "几内亚",
    "GW": "几内亚比绍",
    "KE": "肯尼亚",
    "LS": "莱索托",
    "LR": "利比里亚",
    "LY": "利比亚",
    "MA": "摩洛哥",
    "MO": "澳门",
    "MZ": "莫桑比克",
    "NA": "纳米比亚",
    "NE": "尼日尔",
    "NG": "尼日利亚",
    "RW": "卢旺达",
    "ST": "圣多美和普林西比",
    "SC": "塞舌尔",
    "SL": "塞拉利昂",
    "SO": "索马里",
    "ZA": "南非",
    "SS": "南苏丹",
    "SD": "苏丹",
    "TZ": "坦桑尼亚",
    "TG": "多哥",
    "TN": "突尼斯",
    "UG": "乌干达",
    "ZM": "赞比亚",
    "ZW": "津巴布韦"
}
AMERICA = {
    "AG": "安提瓜和巴布达",
    "AR": "阿根廷",
    "BS": "巴哈马",
    "BB": "巴巴多斯",
    "BZ": "伯利兹",
    "BO": "玻利维亚",
    "BR": "巴西",
    "CA": "加拿大",
    "CL": "智利",
    "CO": "哥伦比亚",
    "CR": "哥斯达黎加",
    "CU": "古巴",
    "DM": "多米尼克",
    "DO": "多米尼加共和国",
    "EC": "厄瓜多尔",
    "SV": "萨尔瓦多",
    "GF": "法属圭亚那",
    "GT": "危地马拉",
    "GY": "圭亚那",
    "HT": "海地",
    "HN": "洪都拉斯",
    "JM": "牙买加",
    "MX": "墨西哥",
    "NI": "尼加拉瓜",
    "PA": "巴拿马",
    "PY": "巴拉圭",
    "PE": "秘鲁",
    "PR": "波多黎各",
    "DO": "多米尼加共和国",
    "SR": "苏里南",
    "TT": "特立尼达和多巴哥",
    "US": "美国",
    "UY": "乌拉圭",
    "VE": "委内瑞拉"
}
EUROPE = {
    "AD": "安道尔",
    "AL": "阿尔巴尼亚",
    "AT": "奥地利",
    "BY": "白俄罗斯",
    "BE": "比利时",
    "BA": "波斯尼亚和黑塞哥维那",
    "BG": "保加利亚",
    "HR": "克罗地亚",
    "CY": "塞浦路斯",
    "CZ": "捷克",
    "DK": "丹麦",
    "EE": "爱沙尼亚",
    "FI": "芬兰",
    "FR": "法国",
    "DE": "德国",
    "GR": "希腊",
    "HU": "匈牙利",
    "IS": "冰岛",
    "IE": "爱尔兰",
    "IT": "意大利",
    "LV": "拉脱维亚",
    "LI": "列支敦士登",
    "LT": "立陶宛",
    "LU": "卢森堡",
    "MK": "北马其顿",
    "MT": "马耳他",
    "MD": "摩尔多瓦",
    "ME": "黑山",
    "NL": "荷兰",
    "NO": "挪威",
    "PL": "波兰",
    "PT": "葡萄牙",
    "RO": "罗马尼亚",
    "RU": "俄罗斯",
    "RS": "塞尔维亚",
    "SK": "斯洛伐克",
    "SI": "斯洛文尼亚",
    "ES": "西班牙",
    "SE": "瑞典",
    "CH": "瑞士",
    "GB": "英国"
}
K_ZONE = {
    "BE": "比利时",
    "IC": "加纳利群岛",
    "FR": "法国",
    "DE": "德国",
    "NL": "荷兰",
    "SM": "圣马力诺",
    "GB": "英国",
    "VA": "梵蒂冈"
}
M_ZONE = {
    "AT": "奥地利",
    "DK": "丹麦",
    "IE": "爱尔兰",
    "LU": "卢森堡",
    "MC": "摩纳哥",
    "SE": "瑞典",
    "CH": "瑞士"
}
E_ZONE = {
    "BG": "保加利亚",
    "HR": "克罗地亚",
    "CY": "塞浦路斯",
    "CZ": "捷克共和国",
    "EE": "爱沙尼亚",
    "FO": "法罗群岛",
    "GL": "格林兰岛",
    "HU": "匈牙利",
    "IS": "冰岛",
    "IL": "以色列",
    "LV": "拉脱维亚",
    "LT": "立陶宛",
    "MT": "马耳他",
    "PL": "波兰",
    "RO": "罗马尼亚",
    "RS": "塞尔维亚",
    "SK": "斯洛伐克共和国",
    "SI": "斯洛文尼亚",
    "TR": "土耳其"
}

country_hippo = {'US': 11987, 'PH': 13410, 'QA': 13411, 'MV': 13412, 'NP': 13413, 'LK': 13414, 'TR': 13415, 'BN': 13416,
                 'IL': 13417, 'ID': 13418, 'AL': 13419, 'EE': 13420, 'AD': 13421, 'AT': 13422, 'BG': 13423, 'BE': 13424,
                 'IS': 13425, 'RU': 13426, 'FI': 13427, 'HR': 13428, 'LU': 13429, 'RO': 13430, 'MC': 13431, 'SI': 13432,
                 'UA': 13433, 'GR': 13434, 'MX': 13435, 'JM': 13436, 'DZ': 13450, 'EG': 13451, 'ET': 13452, 'AO': 13453,
                 'BJ': 13454, 'BW': 13455, 'BF': 13456, 'BI': 13457, 'TG': 13458, 'ER': 13459, 'CV': 13460, 'GP': 13461,
                 'CG': 13462, 'DJ': 13463, 'GN': 13464, 'GW': 13465, 'GH': 13466, 'GA': 13467, 'ZW': 13468, 'CM': 13469,
                 'KM': 13470, 'CI': 13471, 'KE': 13472, 'LS': 13473, 'LR': 13474, 'LY': 13475, 'RW': 13476, 'MG': 13477,
                 'MW': 13478, 'ML': 13479, 'YT': 13480, 'MU': 13481, 'MR': 13482, 'MA': 13483, 'MZ': 13484, 'NA': 13485,
                 'ZA': 13486, 'NE': 13487, 'NG': 13488, 'SL': 13489, 'SN': 13490, 'SC': 13491, 'ST': 13492, 'SZ': 13493,
                 'SD': 13494, 'SO': 13495, 'TZ': 13496, 'TN': 13497, 'UG': 13498, 'IO': 13499, 'ZM': 13500, 'TD': 13501,
                 'GI': 13502, 'MS': 13507, 'NI': 13508, 'NF': 13509, 'SV': 13510, 'LC': 13511, 'TC': 13512, 'TT': 13513,
                 'GT': 13514, 'VI': 13515, 'VG': 13516, 'MQ': 13517, 'CU': 13518, 'HT': 13519, 'HN': 13520, 'CR': 13521,
                 'GD': 13522, 'GL': 13523, 'DM': 13524, 'BM': 13525, 'PA': 13526, 'PR': 13527, 'BZ': 13528, 'AG': 13529,
                 'AR': 13530, 'BO': 13531, 'BR': 13532, 'CO': 13533, 'EC': 13534, 'GY': 13535, 'PY': 13536, 'PE': 13537,
                 'SR': 13538, 'UY': 13539, 'VE': 13540, 'PG': 13541, 'FJ': 13542, 'KI': 13543, 'MH': 13544, 'FM': 13545,
                 'NR': 13546, 'PW': 13547, 'SB': 13548, 'TO': 13549, 'TV': 13550, 'VU': 13551, 'BA': 13552, 'LV': 13553,
                 'LI': 13554, 'SM': 13555, 'HU': 13556, 'ME': 13557, 'XK': 13558, 'LT': 13559, 'MT': 13560, 'CY': 13561,
                 'MD': 13562, 'GB': 12072, 'CA': 12311, 'AU': 12303, 'NZ': 13401, 'DK': 13405, 'SE': 13403, 'NO': 13407,
                 'IE': 13404, 'CH': 13408, 'SA': 13409, 'NL': 13406, 'FR': 12071, 'DE': 12070, 'ES': 12327, 'IT': 13397,
                 'PT': 13396, 'KY': 13399, 'SK': 13398, 'CL': 13394, 'PL': 12069, 'SG': 12306, 'JP': 12305, 'MY': 13402,
                 'KR': 12304, 'CN': 13395}

# glb = dict(globals)
# print(type(glb))
# glblist = list(glb.values())
# print(type(glblist))


countryList_raw = input("请输入国家代码或者区域代码,如 Asia, America, Europe,或者 us, gb :   ").upper()  # 国家代码自动转成大写
weight = int(input("请输入包裹重量(克): "))

if countryList_raw in ("GLOBALS", "ASIA", "OCEANIA", "AFRICA", "AMERICA", "EUROPE", "K_ZONE", "M_ZONE", "E_ZONE"):
    countryList_raw = globals()[countryList_raw]
    countryList = list(dict(countryList_raw).keys())
    quote_list = countryList_raw

else:

    countryList = countryList_raw.split(" ")
    quote_list = GLOBALS

    print(countryList)

wb = Workbook()


def genXs():
    x_nonce = str(uuid.uuid4())  # 生成随机 UUID
    # print("x_nonce:", x_nonce)

    signKey = "980683EF-46C6-47D5-80C1-7B2CB6B2D0BF"

    str_tohash = str(x_nonce) + signKey

    # 生成 x_signature 的方法
    def md5_hash_str(str_tohash):
        # 创建一个 MD5 对象
        md5 = hashlib.md5()
        # 将输入字符串编码为字节类型，因为 update 方法需要字节类型的参数
        input_bytes = str_tohash.encode('utf-8')
        # 使用 update 方法更新 MD5 对象的状态，传入要计算哈希值的字节数据
        md5.update(input_bytes)
        # 获取计算得到的 MD5 哈希值，并以十六进制字符串的形式返回
        return md5.hexdigest()

    # 调用自定义的 md5_hash_str 函数进行哈希计算
    x_signature = str(md5_hash_str(str_tohash))
    # print("hash_result:",x_signature)

    return (x_nonce, x_signature)


# hoobuy 生成类似长度的整数部分
integer_part = random.randint(171446900, 171447000)
# 生成类似长度的小数部分，这里控制小数部分为 9 位
decimal_part = random.randint(100000000, 999999999) / 1e9
# 组合整数部分和小数部分
x_client_id = str(integer_part + decimal_part)
print("x-client-id:", x_client_id)

url_hoobuy = "https://api.hoobuy.com/hoobuy_express/express/pub/postage"

x_nonce, x_signature = genXs();


def countryFee(country, weight):
    # 公共变量 my_feedata 存储全部获取的邮费信息
    # country = input("请输入国家代码: ").upper()  # 国家代码自动转成大写
    # weight = int(input("请输入包裹重量(克): "))

    my_feedata = {}
    my_feedata["国家"] = country
    my_feedata["重量"] = weight
    my_feedata["网站"] = []

    # 输出excel数据
    sheetData = []

    # 1. mule 邮费查询
    url_mule = "https://mulebuy.com/wp-admin/admin-ajax.php?action=get_estimation_query_prices"

    headers = {
        "Content-Type": "multipart/form-data; boundary=----WebKitFormBoundaryLxGfRixPC0oRahOb",
        "Origin": "https://mulebuy.com",
        "Referer": "https://mulebuy.com/zh/estimation/",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
        "Cookie": "PHPSESSID=k8d8da2crlctau78h1qlb3cdvl; wmc_current_currency=USD; wmc_current_currency_old=USD; pll_language=zh; user_temp_intercom=d80184f3df9ff13bb88ed11367c56fe1; name_temp_intercom=TempUser_d80184f3df9ff13bb88ed11367c56fe1; _ga=GA1.1.1728349236.1739253174; _tt_enable_cookie=1; _ttp=-TuUqijACLA3NnN_yK14b5ot-Ls.tt.1; __ukey=81sgb4t9x903; intercom-id-tj75osrf=e445c5c6-8929-40ef-8b42-8e8a824b5e22; intercom-session-tj75osrf=; intercom-device-id-tj75osrf=5626b778-73c8-4564-94aa-8e8ca26b36b2; _ga_QFC95GJ39P=GS1.1.1739254996.2.0.1739254996.60.0.1542245267",
        "accept-language": "zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7,ja-JP;q=0.6,ja;q=0.5",
        "accept": "*/*",
        "Accept-Encoding": "gzip, deflate, br, zstd"
    }

    data = {
        "destination": country,
        "weight": weight,
        "features": None,
        "length": None,
        "width": None,
        "height": None,
        "username": None,
        "password": None,
        "terms": 1

    }
    # response = requests.post(url=url_mule, data=data,headers=headers, params=params_mule)
    try:
        response = requests.post(url=url_mule, data=data, verify=False)

        # 接口返回json数据
        fee_datas = response.json()

        curSiteInfo = {"siteName": "mulebuy", "venders": []}
        # 将 mule 数据存入 my_feedata
        for fee in fee_datas['data']:
            if fee['available']:
                # print(fee)
                curSiteInfo['venders'].append(
                    {"venderName": fee['name'],
                     '总价': fee['feeDetail']['total'],
                     '首重价格': fee['feeDetail']['feeFirst'],
                     "额外重量价格": fee['feeDetail']["feeContinue"],
                     "操作费": fee['feeDetail']["operationFee"],
                     "服务费": fee['feeDetail']["serviceFee"],
                     "最低重量限制": fee['restrictions']["minWeight"],
                     "最高重量限制": fee['restrictions']["maxWeight"],
                     "尺寸限制": fee['restrictions']["dimensionRestriction"],
                     "体积重量计费规则": fee['restrictions']["volumeWeightRule"],
                     "运输时间": fee["transitTime"]},
                )

        for item in curSiteInfo['venders']:
            print(item)
            sheetData.append(
                (country,
                 weight,
                 curSiteInfo['siteName'],
                 item['venderName'],
                 item['总价'],
                 item['首重价格'],
                 item['额外重量价格'],
                 item['操作费'],
                 item['服务费'],
                 item['最低重量限制'],
                 item['最高重量限制'],
                 item['尺寸限制'],
                 item['体积重量计费规则'],
                 item['运输时间'],
                 )
            )

        my_feedata["网站"].append(curSiteInfo)
        print("mule 查询完毕")
    except Exception as e:
        print("mule 请求错误")
        print(e)
    # print(response.status_code)
    # print(response.json())

    # 2 获取orientdig 邮费信息
    url_orientdig = "https://orientdig.com/wp-admin/admin-ajax.php?action=get_estimation_query_prices"

    headers = {
        "Content-Type": "multipart/form-data; boundary=----WebKitFormBoundarywQkmSLhKNWpKOJ9u",
        "accept": "*/*",
        "Accept-Encoding": "gzip, deflate, br, zstd",
        # "Origin": "https://cnfans.com",
        # "Referer": "https://cnfans.com/zh/estimation/",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
        # "Cookie": "PHPSESSID=ohp6s0j3vqfcnrbrr1vatq527u; wmc_current_currency=USD; wmc_current_currency_old=USD",
        "accept-language": "zh-CN,zh;q=0.9",

    }
    data = {
        "destination": country,
        "weight": weight,
        "terms": 1

    }

    try:
        response = requests.post(url=url_orientdig, data=data, verify=False)
        fee_datas = response.json()
        # 提取邮费信息

        curSiteInfo = {"siteName": "orientDig", "venders": []}
        for fee in fee_datas['data']:
            if fee['available']:
                curSiteInfo['venders'].append(
                    {"venderName": fee['name'],
                     '总价': fee['feeDetail']['total'],
                     '首重价格': fee['feeDetail']['feeFirst'],
                     "额外重量价格": fee['feeDetail']["feeContinue"],
                     "操作费": fee['feeDetail']["operationFee"],
                     "服务费": fee['feeDetail']["serviceFee"],
                     "最低重量限制": fee['restrictions']["minWeight"],
                     "最高重量限制": fee['restrictions']["maxWeight"],
                     "尺寸限制": fee['restrictions']["dimensionRestriction"],
                     "体积重量计费规则": fee['restrictions']["volumeWeightRule"],
                     "运输时间": fee["transitTime"]},
                )

        for item in curSiteInfo['venders']:
            print(item)
            sheetData.append(
                (country,
                 weight,
                 curSiteInfo['siteName'],
                 item['venderName'],
                 item['总价'],
                 item['首重价格'],
                 item['额外重量价格'],
                 item['操作费'],
                 item['服务费'],
                 item['最低重量限制'],
                 item['最高重量限制'],
                 item['尺寸限制'],
                 item['体积重量计费规则'],
                 item['运输时间'],
                 )
            )
        my_feedata["网站"].append(curSiteInfo)
        print("orientdig 查询完毕")
    except Exception as e:
        print("获取orientdig 信息错误")
        print(e)

    # 3 hoobuy 邮费信息
    # def genXs():
    #     x_nonce = str(uuid.uuid4())  # 生成随机 UUID
    #     # print("x_nonce:", x_nonce)
    #
    #     signKey = "980683EF-46C6-47D5-80C1-7B2CB6B2D0BF"
    #
    #     str_tohash = str(x_nonce) + signKey
    #
    #     # 生成 x_signature 的方法
    #     def md5_hash_str(str_tohash):
    #         # 创建一个 MD5 对象
    #         md5 = hashlib.md5()
    #         # 将输入字符串编码为字节类型，因为 update 方法需要字节类型的参数
    #         input_bytes = str_tohash.encode('utf-8')
    #         # 使用 update 方法更新 MD5 对象的状态，传入要计算哈希值的字节数据
    #         md5.update(input_bytes)
    #         # 获取计算得到的 MD5 哈希值，并以十六进制字符串的形式返回
    #         return md5.hexdigest()
    #
    #     # 调用自定义的 md5_hash_str 函数进行哈希计算
    #     x_signature = str(md5_hash_str(str_tohash))
    #     # print("hash_result:",x_signature)
    #
    #     return (x_nonce, x_signature)
    #

    url_hoobuy = "https://api.hoobuy.com/hoobuy_express/express/pub/postage"

    x_nonce, x_signature = genXs();

    headers = {
        "x-client-id": x_client_id,
        "x-nonce": x_nonce,
        "x-signature": x_signature,
        "x-platform": "web",
        "x-version": "hoobuy-production-web-1.0",
        "Origin": "https://hoobuy.com",
        # "Referer":"https://cnfans.com/zh/estimation/",
        # "priority":"u=1, i",
        # "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/133.0.0.0 Safari/537.36",
        # "Cookie": "pll_language=zh; lf_session_id=bcc81373-fc65-40d1-ad49-2e5389a55736; lf_first_visit=1739070628072; lf_prev_visit=1739070628072; __ukey=81n288jdx800; _gcl_au=1.1.997165745.1739070628; _tt_enable_cookie=1; _ttp=9ZfJV76oBGK2r66PT-bk81I6gAS.tt.1; _gid=GA1.2.95782528.1739070629; _ss_s_uid=794d38360ed3b348b16b30f5ba05689f; _hjSessionUser_3640651=eyJpZCI6ImMyMTM2OTdkLWM3YWQtNWVhMi1hNjQ4LTcwZmI0YWRkYjRiZSIsImNyZWF0ZWQiOjE3MzkwNzA2Mjg0ODUsImV4aXN0aW5nIjp0cnVlfQ==; wmc_current_currency=USD; wmc_current_currency_old=USD; cf_clearance=0RyHZJKQqkCkf.6lgMzYGOP4ICl_sW7DSVUWJE4SbKA-1739081585-1.2.1.1-yy4UsEorkA9CYi1taG3wfrsgtcpLkRpgUbhZJfgIMSueiYIadPVbCQg0dWnx8p8kDwnTGBSI.ZdvPGnuck8N8TT4xLcmqRwBcnyCwcPUISB6Mkkzm65lTt7gJN2my56VKkq60HTuP8_hCuBMwEUOD6VL6Y9pdK_iz3i.oKTZHmp64ZKyulIuS3jkSEDpP6V0IU5iPSxLBmVZRHA5Chdv6s1QoSouaVjjWfcaa9vf_Bdt3ijlhyLIQxQvwmhQujg.ZxLaGyoohbeOJ7YdZ3kNg5Rb5qW.s4F1te1npEBHoEM; lf_this_visit=1739081657585; lf_session_count=2; _hjSession_3640651=eyJpZCI6IjljOTEzZjRjLTY1YTctNDUwNy1hZDQzLTEyOWQ2MWYwNzFlNSIsImMiOjE3MzkwODE2NTc2MzIsInMiOjAsInIiOjAsInNiIjowLCJzciI6MCwic2UiOjAsImZzIjowLCJzcCI6MH0=;"
        #           " _ga=GA1.2.1556063410.1739070628; _gat_gtag_UA_257386089_1=1; lf_prev_send_time=1739082046994; _ga_GVNMMZMPG4=GS1.1.1739081657.2.1.1739082052.51.0.0",
        "Accept": "*application/json",
        # "Accept-encoding":"gzip, deflate, br, zstd",
        # # "accept-language": "zh-CN,zh;q=0.9"
        # "Connection":"keep-alive"

    }
    data = {"country": country, "requestType": 1,
            "arr": [{"weight": weight, "long": 0, "width": 0, "height": 0, "count": 1, "coefficient": 1}]}
    try:
        response = requests.post(url=url_hoobuy, json=data, headers=headers, verify=False)

        fee_datas = response.json()

        # print("response:", fee_datas)

        curSiteInfo = {"siteName": "hoobuy", "venders": []}

        ids = []
        for fee in fee_datas['data']['lines']:
            if fee['state'] == "available":
                ids.append(fee['id'])

        for fee in fee_datas['data']['lines']:
            if fee['state'] == "available":
                curSiteInfo['venders'].append(
                    {"id": fee['id'],
                     "venderName": fee['lineName'],
                     '总价': str(float(fee['operationFee']) + float(fee['price'])),
                     '首重价格': 0,  #
                     "额外重量价格": 0,  #
                     "操作费": fee["operationFee"],
                     "服务费": 0,  #
                     "最低重量限制": None,  #
                     "最高重量限制": None,  #
                     "尺寸限制": None,  #
                     "体积重量计费规则": None,  #
                     "运输时间": fee["timeRequired"]},
                )

        print("ids: ", ids)
        details = []
        for line_id in ids:
            x_nonce, x_signature = genXs();

            headers = {
                "x-client-id": x_client_id,
                "x-nonce": x_nonce,
                "x-signature": x_signature,
                "x-platform": "web",
                "x-version": "hoobuy-production-web-1.0",
                "Origin": "https://hoobuy.com",
                # "Referer":"https://cnfans.com/zh/estimation/",
                # "priority":"u=1, i",
                # "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/133.0.0.0 Safari/537.36",
                # "Cookie": "pll_language=zh; lf_session_id=bcc81373-fc65-40d1-ad49-2e5389a55736; lf_first_visit=1739070628072; lf_prev_visit=1739070628072; __ukey=81n288jdx800; _gcl_au=1.1.997165745.1739070628; _tt_enable_cookie=1; _ttp=9ZfJV76oBGK2r66PT-bk81I6gAS.tt.1; _gid=GA1.2.95782528.1739070629; _ss_s_uid=794d38360ed3b348b16b30f5ba05689f; _hjSessionUser_3640651=eyJpZCI6ImMyMTM2OTdkLWM3YWQtNWVhMi1hNjQ4LTcwZmI0YWRkYjRiZSIsImNyZWF0ZWQiOjE3MzkwNzA2Mjg0ODUsImV4aXN0aW5nIjp0cnVlfQ==; wmc_current_currency=USD; wmc_current_currency_old=USD; cf_clearance=0RyHZJKQqkCkf.6lgMzYGOP4ICl_sW7DSVUWJE4SbKA-1739081585-1.2.1.1-yy4UsEorkA9CYi1taG3wfrsgtcpLkRpgUbhZJfgIMSueiYIadPVbCQg0dWnx8p8kDwnTGBSI.ZdvPGnuck8N8TT4xLcmqRwBcnyCwcPUISB6Mkkzm65lTt7gJN2my56VKkq60HTuP8_hCuBMwEUOD6VL6Y9pdK_iz3i.oKTZHmp64ZKyulIuS3jkSEDpP6V0IU5iPSxLBmVZRHA5Chdv6s1QoSouaVjjWfcaa9vf_Bdt3ijlhyLIQxQvwmhQujg.ZxLaGyoohbeOJ7YdZ3kNg5Rb5qW.s4F1te1npEBHoEM; lf_this_visit=1739081657585; lf_session_count=2; _hjSession_3640651=eyJpZCI6IjljOTEzZjRjLTY1YTctNDUwNy1hZDQzLTEyOWQ2MWYwNzFlNSIsImMiOjE3MzkwODE2NTc2MzIsInMiOjAsInIiOjAsInNiIjowLCJzciI6MCwic2UiOjAsImZzIjowLCJzcCI6MH0=;"
                #           " _ga=GA1.2.1556063410.1739070628; _gat_gtag_UA_257386089_1=1; lf_prev_send_time=1739082046994; _ga_GVNMMZMPG4=GS1.1.1739081657.2.1.1739082052.51.0.0",
                "Accept": "*application/json",
                # "Accept-encoding":"gzip, deflate, br, zstd",
                # # "accept-language": "zh-CN,zh;q=0.9"
                # "Connection":"keep-alive"

            }
            data = {"lineId": line_id}

            url_detail = 'https://api.hoobuy.com/hoobuy_order/pub/v2/express/line/detail'

            response = requests.post(url=url_detail, json=data, headers=headers, verify=False)
            detail = response.json()
            details.append({'id': line_id,
                            "最低重量限制": detail['data']['weightLimit']['lowestWeight'],
                            "最高重量限制": detail['data']['weightLimit']['highestWeight'],
                            "尺寸限制": str(detail['data']['sizeLimit'])
                            })
            # time.sleep(0.5)

        merged_list = []

        for dic1, dic2 in zip(curSiteInfo['venders'], details):
            merged_dic = dic1.copy()
            merged_dic.update(dic2)
            merged_list.append(merged_dic)

        curSiteInfo['venders'] = merged_list

        for item in curSiteInfo['venders']:
            print(item)
            sheetData.append(
                (country,
                 weight,
                 curSiteInfo['siteName'],
                 item['venderName'],
                 item['总价'],
                 item['首重价格'],
                 item['额外重量价格'],
                 item['操作费'],
                 item['服务费'],
                 item['最低重量限制'],
                 item['最高重量限制'],
                 item['尺寸限制'],
                 item['体积重量计费规则'],
                 item['运输时间'],
                 )
            )

        my_feedata["网站"].append(curSiteInfo)
        print("hoobuy 查询完毕")
    except Exception as e:
        print(e)
        print("hoobuy 查询错误")

    # 4 joyabuy 邮费信息
    boundary = f"----WebKitFormBoundary{uuid.uuid4().hex[:16]}"
    multipart_data = (
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="destination"\r\n\r\n'
        f"{country}\r\n"
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="weight"\r\n\r\n'
        f"{weight}\r\n"
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="features"\r\n\r\n'
        f"\r\n"
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="length"\r\n\r\n'
        f"\r\n"
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="width"\r\n\r\n'
        f"\r\n"
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="height"\r\n\r\n'
        f"\r\n"
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="username"\r\n\r\n'
        f"\r\n"
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="password"\r\n\r\n'
        f"\r\n"
        f"--{boundary}--"
    )

    url_joya = "https://joyabuy.com/wp-admin/admin-ajax.php?action=get_estimation_query_prices"

    params = {
        "action": "get_estimation_query_prices"
    }

    # 目标 URL
    url_joya = "https://joyabuy.com/wp-admin/admin-ajax.php?action=get_estimation_query_prices"

    headers_joya = {
        # ":method": "POST",
        # ":authority": "joyabuy.com",
        # ":scheme": "https",
        # ":path": "/wp-admin/admin-ajax.php?action=get_estimation_query_prices",
        # "content-length": str(len(data)),
        "pragma": "no-cache",
        "cache-control": "no-cache",
        "sec-ch-ua-platform": '"Windows"',
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/133.0.0.0 Safari/537.36",
        # "sec-ch-ua": '"Not(A:Brand";v="99", "Google Chrome";v="133", "Chromium";v="133")',
        "content-type": f"multipart/form-data; boundary={boundary}",
        # "content-type": multipart_data.content_type,
        # "sec-ch-ua-mobile": "?0",
        "Accept": "*/*",
        "origin": "https://joyabuy.com",
        "Referer": "https://joyabuy.com/estimation/",
        # "sec-fetch-site": "same-origin",
        # "sec-fetch-mode": "cors",
        # "sec-fetch-dest": "empty",
        "Accept-Encoding": "gzip, deflate, br, zstd",
        "Accept-Language": "zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7,ja-JP;q=0.6,ja;q=0.5",
        # "priority": "u=1, i",
        # "Cookie": "wmc_current_currency=USD; wmc_current_currency_old=USD; pll_language=en; _ss_s_uid=9eb6242c0bcb35c9de9555eb2b9c731b; _ga=GA1.1.1778277922.1739259791; __ukey=81sml3atx5; PHPSESSID=jmfohumqqujm5v4qkfvpqcreqo; _ga_LLNZ3BEEWR=GS1.1.1739324439.5.0.1739324439.0.0.0"
    }

    # with httpx.Client(http2=True) as client:
    #     client.headers.update(headers_joya)
    #     response = client.post(url=url_joya, content=multipart_data.encode())
    #     request = response.request
    #     # 打印请求体
    #     # print("request body:", request.content.decode())

    try:
        response = requests.post(url=url_joya, data=multipart_data.encode(), headers=headers_joya, verify=False)

        fee_datas = response.json()

        curSiteInfo = {"siteName": "joyabuy", "venders": []}

        for fee in fee_datas['data']:
            if fee['available']:
                curSiteInfo['venders'].append(
                    {"venderName": fee['name'],
                     '总价': fee['feeDetail']['total'],
                     '首重价格': fee['feeDetail']['feeFirst'],
                     "额外重量价格": fee['feeDetail']["feeContinue"],
                     "操作费": fee['feeDetail']["operationFee"],
                     "服务费": fee['feeDetail']["serviceFee"],
                     "最低重量限制": fee['restrictions']["minWeight"],
                     "最高重量限制": fee['restrictions']["maxWeight"],
                     "尺寸限制": fee['restrictions']["dimensionRestriction"],
                     "体积重量计费规则": fee['restrictions']["volumeWeightRule"],
                     "运输时间": fee["transitTime"]},
                )

        # 打印所以有邮费信息

        for item in curSiteInfo['venders']:
            print(item)
            sheetData.append(
                (country,
                 weight,
                 curSiteInfo['siteName'],
                 item['venderName'],
                 item['总价'],
                 item['首重价格'],
                 item['额外重量价格'],
                 item['操作费'],
                 item['服务费'],
                 item['最低重量限制'],
                 item['最高重量限制'],
                 item['尺寸限制'],
                 item['体积重量计费规则'],
                 item['运输时间'],
                 )
            )

        my_feedata['网站'].append(curSiteInfo)
        print("joyabuy 查询完毕")
    except Exception as e:
        print(e)
        print("joyabuy 查询错误")

    # 5 acbuy
    url_acbuy = "https://acbuy.com/prefix-api/store-logistics/estimate/route"

    headers = {
        "accept-language": "zh-CN,zh;q=0.9"
    }
    data = {"countryCode": country, "weight": weight, "length": "", "width": "", "height": "", "itemLimitList": [],
            "sortBy": "price", "sort": "asc"}

    try:
        response = requests.post(url=url_acbuy, json=data, verify=False)

        # print(response.status_code)
        # print(response.json())

        fee_datas = response.json()

        curSiteInfo = {"siteName": "acbuy", "venders": []}

        for fee in fee_datas['data']:
            feeMap = fee['feeDetail']['feeCompositions']['feeMap']
            # print(type(feeMap))
            # print(feeMap['A1'])
            try:
                feeMap_dict = json.loads(feeMap['A1'])
                continuedPrice = feeMap_dict[0]['continuedPrice']
                firstPrice = feeMap_dict[0]['firstPrice']
                weightRangeStart = feeMap_dict[0]['weightRangeStart']
                weightRangeEnd = feeMap_dict[0]['weightRangeEnd']
            except:
                continuedPrice = None
                firstPrice = None
                weightRangeStart = None
                weightRangeEnd = None

            # firstPrice = feeMap_dict[0]['firstPrice']
            # print(fee['feeDetail']["feeDTOList"][1]["feeAmount"],"操作费")

            start_date = fee["carrierProductLineInfo"]["fastestDuration"]
            end_date = fee["carrierProductLineInfo"]["slowestDuration"]
            duration = str(start_date) + "-" + str(end_date)

            curSiteInfo['venders'].append(
                {"venderName": fee['carrierProductCode'],
                 '总价': fee['feeDetail']['totalAmount'],
                 '首重价格': firstPrice,
                 "额外重量价格": continuedPrice,
                 "操作费": fee['feeDetail']["feeDTOList"][1]["feeAmount"],
                 "服务费": 0,
                 "最低重量限制": weightRangeStart,
                 "最高重量限制": weightRangeEnd,
                 "尺寸限制": None,
                 "体积重量计费规则": None,
                 "运输时间": duration},

            )

        for item in curSiteInfo['venders']:
            print(item)
            sheetData.append(
                (country,
                 weight,
                 curSiteInfo['siteName'],
                 item['venderName'],
                 item['总价'],
                 item['首重价格'],
                 item['额外重量价格'],
                 item['操作费'],
                 item['服务费'],
                 item['最低重量限制'],
                 item['最高重量限制'],
                 item['尺寸限制'],
                 item['体积重量计费规则'],
                 item['运输时间'],
                 )
            )

        my_feedata['网站'].append(curSiteInfo)
        print("acbuy 查询完毕")
    except Exception as e:
        print(e)
        print("acbuy 查询错误")

    # 6, cnfans

    # 创建会话，模拟 Chrome
    session = tls_client.Session(
        client_identifier="chrome_133",  # 伪装 Chrome 120
        random_tls_extension_order=True  # 让 TLS 扩展字段顺序随机化
    )

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    }

    # 目标 URL
    url_cfans = "https://cnfans.com/wp-admin/admin-ajax.php?action=get_estimation_query_prices"

    # boundary = f"----WebKitFormBoundary{uuid.uuid4().hex[:16]}"
    boundary = f"----WebKitFormBoundary{uuid.uuid4().hex}"
    multipart_data = (
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="destination"\r\n\r\n'
        f"{country}\r\n"
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="weight"\r\n\r\n'
        f"{weight}\r\n"
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="features"\r\n\r\n'
        f"\r\n"
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="length"\r\n\r\n'
        f"\r\n"
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="width"\r\n\r\n'
        f"\r\n"
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="height"\r\n\r\n'
        f"\r\n"
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="username"\r\n\r\n'
        f"\r\n"
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="password"\r\n\r\n'
        f"\r\n"
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="terms"\r\n\r\n'
        f"1\r\n"
        f"--{boundary}--\r\n"
    )

    headers_cfans = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/133.0.0.0 Safari/537.36",
        "sec-ch-ua": '"Not(A:Brand";v="99", "Google Chrome";v="133", "Chromium";v="133")',
        "Content-Type": f"multipart/form-data; boundary={boundary}",
        "accept": "*/*",
        "origin": "https://cnfans.com",
        "referer": "https://cnfans.com/zh/estimation/",
        "accept-encoding": "gzip, deflate, br, zstd",
        "accept-language": "zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7,ja-JP;q=0.6,ja;q=0.5",
    }

    try:

        response = session.post(url_cfans, headers=headers_cfans, data=multipart_data.encode())

        fee_datas = response.json()

        curSiteInfo = {"siteName": "cnfans", "venders": []}

        for fee in fee_datas['data']:
            if fee['available']:
                curSiteInfo['venders'].append(
                    {"venderName": fee['name'],
                     '总价': fee['feeDetail']['total'],
                     '首重价格': fee['feeDetail']['feeFirst'],
                     "额外重量价格": fee['feeDetail']["feeContinue"],
                     "操作费": fee['feeDetail']["operationFee"],
                     "服务费": fee['feeDetail']["serviceFee"],
                     "最低重量限制": fee['restrictions']["minWeight"],
                     "最高重量限制": fee['restrictions']["maxWeight"],
                     "尺寸限制": fee['restrictions']["dimensionRestriction"],
                     "体积重量计费规则": fee['restrictions']["volumeWeightRule"],
                     "运输时间": fee["transitTime"]},
                )

        # 打印所以有邮费信息

        for item in curSiteInfo['venders']:
            print(item)
            sheetData.append(
                (country,
                 weight,
                 curSiteInfo['siteName'],
                 item['venderName'],
                 item['总价'],
                 item['首重价格'],
                 item['额外重量价格'],
                 item['操作费'],
                 item['服务费'],
                 item['最低重量限制'],
                 item['最高重量限制'],
                 item['尺寸限制'],
                 item['体积重量计费规则'],
                 item['运输时间'],
                 )
            )

        my_feedata['网站'].append(curSiteInfo)
        print("cnfans 查询完毕")
    except Exception as e:
        print(e)
        print("cnfans 查询错误")

    # 7 hippobuy

    url_hippo = "https://api-jiyun-v3.haiouoms.com/api/client/express/price-query"
    headers = {
        "Host": "api-jiyun-v3.haiouoms.com",
        "Connection": "keep-alive",
        "Content-Length": "150",
        "language": "zh_CN",
        "sec-ch-ua-platform": "\"Windows\"",
        "Authorization": "None",
        "sec-ch-ua": "\"Not(A:Brand\";v=\"99\", \"Google Chrome\";v=\"133\", \"Chromium\";v=\"133\"",
        "currency": "USD",
        "sec-ch-ua-mobile": "?0",
        "App-key": "arIM1n8tmcrHgD1jHz0Zt8H1XPxUjVen",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/133.0.0.0 Safari/537.36",
        "Accept": "application/json, text/plain, */*",
        "Content-Type": "application/json;charset=UTF-8",
        "X-Uuid": "dd70dbad-9df5-4813-b17d-910a60dd54ef",
        "Origin": "https://hippoobuy.com",
        "Sec-Fetch-Site": "cross-site",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Dest": "empty",
        "Referer": "https://hippoobuy.com/",
        "Accept-Encoding": "gzip, deflate, br, zstd",
        "Accept-Language": "zh-CN,zh;q=0.9"
    }

    if country in country_hippo.keys():
        data = {"warehouse_id": 1250, "country_id": country_hippo[country], "area_id": "", "sub_area_id": "",
                "weight": weight * 1000,
                "length": "", "width": "", "height": "", "prop_ids": [], "postcode": ""}

        try:

            response = requests.post(url=url_hippo, headers=headers, json=data, verify=False)
            fee_datas = response.json()
            curSiteInfo = {"siteName": "hippobuy", "venders": []}

            for fee in fee_datas['data']:
                curSiteInfo['venders'].append(
                    # {"venderName": fee['name']},
                    # {'总价': fee['feeDetail']['total']},
                    # {'首重价格': fee['first_money']},
                    # {"额外重量价格": fee['feeDetail']["feeContinue"]},
                    # {"操作费": fee['feeDetail']["operationFee"]},
                    # {"服务费": fee['feeDetail']["serviceFee"]},
                    # {"最低重量限制": fee['min_weight']},
                    # {"最高重量限制": fee['max_weight'],
                    # {"尺寸限制": fee['restrictions']["dimensionRestriction"]},
                    # {"体积重量计费规则": fee['restrictions']["volumeWeightRule"]},
                    # {"运输时间": fee["reference_time"]}

                    {"venderName": fee['cn_name'],
                     '总价': fee['count_first'] / 100,
                     '首重价格': fee['first_money'] / 100,
                     "额外重量价格": None,
                     "操作费": None,
                     "服务费": None,
                     "最低重量限制": fee['min_weight'] / 1000,
                     "最高重量限制": fee['max_weight'] / 1000,
                     "尺寸限制": None,
                     "体积重量计费规则": fee['remark'],
                     "运输时间": fee["reference_time"]},
                )
                # print(fee)
            for item in curSiteInfo['venders']:
                print(item)
                sheetData.append(
                    (country,
                     weight,
                     curSiteInfo['siteName'],
                     item['venderName'],
                     item['总价'],
                     item['首重价格'],
                     item['额外重量价格'],
                     item['操作费'],
                     item['服务费'],
                     item['最低重量限制'],
                     item['最高重量限制'],
                     item['尺寸限制'],
                     item['体积重量计费规则'],
                     item['运输时间'],
                     )
                )
            my_feedata['网站'].append(curSiteInfo)
            print("hippobuy 查询完毕")

        except Exception as e:
            print(e)
            print("hippobuy 查询错误")

        print(sheetData)

    # 创建一个新的工作簿
    # wb = Workbook()

    # 获取活动工作表
    # ws = wb.active
    ws = wb.create_sheet(f"{quote_list[country]}-运费表")

    # 设置工作表标题
    # ws.title = f"{country}-运费表"

    # 写入数据到单元格
    # 写入数据到单元格
    ws['A1'] = '目的国家'
    ws['B1'] = '重量'
    ws['C1'] = '网站'
    ws['D1'] = '承运公司'
    ws['E1'] = '总价'
    ws['F1'] = '首重价格'
    ws['G1'] = '额外重量价格'
    ws['H1'] = '操作费'
    ws['I1'] = '服务费'
    ws['J1'] = '最低重量限制'
    ws['K1'] = '最高重量限制'
    ws['L1'] = '尺寸限制'
    ws['M1'] = '体积重量计费规则'
    ws['N1'] = '运输时间'

    # 写入多行数据

    for row in sheetData:
        ws.append(row)

    max_rows = ws.max_row  # 获取最大行
    max_columns = ws.max_column
    align = Alignment(horizontal='center', vertical='center')
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 24

    for i in range(1, max_rows + 1):
        for j in range(1, max_columns + 1):
            if i > 1 and (j == 4 or j == 12 or j == 13):
                continue
            ws.cell(i, j).alignment = align

    # 创建一个粗体字体样式
    bold_font = Font(bold=True)
    # 将粗体字体样式应用到 B1 单元格

    for i in range(1, max_rows + 1):
        ws.cell(1, i).font = bold_font

    # for j in range(1,max_columns+1):
    #     ws.column_dimensions[j].width = 15

    # 保存工作簿
    # 将时间戳转换为本地时间元组
    # timestamp=time.time()
    # local_time = time.localtime(timestamp)
    #
    # # 获取日期时间
    # formatted_time = time.strftime("%Y-%m-%d_%H%M%S", local_time)
    # wb.save(f'mailFee_{formatted_time}.xlsx')


for country in countryList:
    countryFee(country, weight)
# 保存工作簿
# 将时间戳转换为本地时间元组
timestamp = time.time()
local_time = time.localtime(timestamp)

# 获取日期时间
formatted_time = time.strftime("%Y-%m-%d_%H%M%S", local_time)

wb.remove(wb["Sheet"])

wb.save(f'mailFee_{formatted_time}.xlsx')

# pyinstaller -F -c --add-binary lib;.  mailfeeAll.py

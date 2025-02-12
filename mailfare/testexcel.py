from openpyxl import Workbook

# 创建一个新的工作簿
wb = Workbook()

# 获取活动工作表
ws = wb.active

# 设置工作表标题
ws.title = "Sheet1"

# 写入数据到单元格
ws['A1'] = '目的国家'
ws['B1'] = '重量'
ws['C1'] = '网站'
ws['D1'] = '承运公司'
ws['E1'] = '总价'
ws['F1'] = '首重价格'
ws['G1'] = '额外重量价格'
ws['H1'] = '操作费'
ws['I1'] = '服务费'
ws['J1'] = '最低重量限制'
ws['K1'] = '最高重量限制'
ws['L1'] = '尺寸限制'
ws['M1'] = '体积重量计费规则'
ws['N1'] = '运输时间'

# 写入多行数据
data = [
    ('张三', 25, '北京'),
    ('李四', 30, '上海'),
    ('王五', 22, '广州')
]

for row in data:
    ws.append(row)

# 保存工作簿
wb.save('new_excel_file.xlsx')
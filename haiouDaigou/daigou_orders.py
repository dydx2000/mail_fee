# 导入库
import time

import requests

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

session = requests.session()

# 1 获取总数

headers = {
    "Content-Type": "application/x-www-form-urlencoded;charset=UTF-8",
    "Origin": "https://jiyun-v4.haiouoms.com",
    "Referer": "https://jiyun-v4.haiouoms.com/",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36",
    "accept-language": "zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7,ja-JP;q=0.6,ja;q=0.5",
    "accept": "application/json, text/plain, */*",
    "language": "zh_CN",
    "authorization": "bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJodHRwczovL2FwaS5oYWlvdW9tcy5jb20vYXBpL2FkbWluL2xvZ2luIiwiaWF0IjoxNzQyMjA1MjAwLCJleHAiOjE3NDQ3OTcyMDAsIm5iZiI6MTc0MjIwNTIwMCwianRpIjoicWV4MVJZQkdZVE5ibDB6dSIsInN1YiI6IjQ0MDQiLCJwcnYiOiJkZjg4M2RiOTdiZDA1ZWY4ZmY4NTA4MmQ2ODZjNDVlODMyZTU5M2E5Iiwicm9sZSI6ImFkbWluIn0.1Jnld6HctTKjCGAlyToXDmpOaJWVmk35p_0dyuC7dcU"

    # "Accept-Encoding": "gzip, deflate, br, zstd"
}

data_count = {
    "platform": None,
    "begin_date": None,
    "end_date": None,
    "date_type": "created_at"
}

url_count = "https://api-jiyun-v3.haiouoms.com/api/admin/daigou-orders/status/count?platform=&begin_date=&end_date=&date_type=created_at"

session.headers = headers

# response = session.get(url=url_count, data=data_count)
# print(response.status_code)
# res = response.json()
# totalPage = res['data']['all']
# print("记录总数为： ", totalPage)

# 2查询主订单数据
# 2.1 获取第一页数据


page = int(input("请输入要查询的页码："))
size = int(input("请输入每页查询的条数："))

url_datas = f"https://api-jiyun-v3.haiouoms.com/api/admin/daigou-orders?page={page}&total=0&size={size}&keyword=&platform=&begin_date=&end_date="
data_orders = {
    "page": page,
    "total": 0,
    "size": size,
    "keyword": None,
    "platform": None,
    "begin_date": None,
    "end_date": None
}

response = session.get(url=url_datas, data=data_orders)

res = response.json()

res_detail = session.get(url="https://api-jiyun-v3.haiouoms.com/api/admin/daigou-orders/89387")

rows = []
sub_rows = []

for data in res['data']:
    # 遍历本页所有数据

    # 获取详情id
    # order_id = data['id']
    # response = session.get(url=f"https://api-jiyun-v3.haiouoms.com/api/admin/daigou-orders/{order_id}")
    # res_detail = response.json()
    #

    if data['skus'] == []:
        jiyun_packageNO = None
        sub_orders = []
    else:
        noList = ''
        sub_orders = []
        for item in data['skus']:
            noList += (item['code'] + ",")
            sub_orders.append(item['code'])

        jiyun_packageNO = noList

    if data['purchase_packages'] == []:
        wuliu_order = None
    else:
        wuliu_order = data['purchase_packages'][0]['express_num']

    if data['purchaser'] == None:
        purchaser = None
    else:
        purchaser = data['purchaser']['username']

    if data['transaction'] == []:
        pay_order = None
        pay_method = None
        pay_amount = None
    else:
        pay_order = data['transaction'][0]['serial_no']
        pay_method = data['transaction'][0]['pay_name']
        pay_amount = int(data['transaction'][0]['amount']) / 100


    print(data['order_sn'])
    if data['warehouse'] is not None:
        whare_house=data['warehouse']['warehouse_name']
    else:
        whare_house = ''


    row = {"order_sn": data['order_sn'],
           "status_name": data['status_name'],
           "country_name": data['address']['country_name'],
           "created_time": data['created_at'],
           "whare_house": whare_house,
           "jiyun_packageNO": jiyun_packageNO,
           "platform_orderNo": data['platform_order_sn'],
           "wuliu_order": wuliu_order,  # data['purchase_packages'][0]['express_num'],   # 就是一个
           "order_user": data['user']['name'],
           "purchaser": purchaser,  # data['purchaser']['username'], # 可能为空
           "order_platform": data['skus'][0]['platform'],
           "address": "",  # 其他接口,
           "pay_status": data["pay_status"],
           "pay_order": pay_order,  # data['transaction'][0]['serial_no'],  # transction [] 可能空数据
           "pay_time": data['paid_at'],
           "pay_method": pay_method,  # data['transaction'][0]['pay_name'],
           "pay_amount": pay_amount,  # data['transaction'][0]['amount'],  # 还要计算?
           "delivery_fee": data['freight_fee'],
           "total_payment": data['amount']
           }
    for sku in data['skus']:
        if sku['name_cn'] is not None:
            sub_name = sku['name_cn']
        else:
            sub_name = sku['name']

        sku_spec = ""
        try:
            # if sku['sku_info']['specs_cn'] is not None:
            if 'specs' in sku['sku_info'].keys():
                for info in sku['sku_info']['specs']:
                    if not (info['label'] is None or info['value'] is None):
                        sku_spec += (info['label'] + ": " + info['value'] + ", ")
                    else:
                        sku_spec = "null"

            elif 'specs_cn' in sku['sku_info'].keys():
                # sku_spec = str(sku['sku_info']['specs_cn'])
                if sku['sku_info']['specs_cn'] == []:
                    sku_spec = 'empty'
                else:
                    for info in sku['sku_info']['specs_cn']:
                        if not (info['label'] is None or info['value'] is None):
                            sku_spec += (info['label'] + ": " + info['value'] + ", ")
                        else:
                            sku_spec = "null"
            else:
                sku_spec = "not exist"
        except Exception as e:
            print("keyError")
            print(e)
            sku_spec = "keyError"

        # 主订单交易状态和子订单状态的逻辑关系
        if data['status_name'] in ['采购中', '待付款', '待入库','已付款','已取消','已入库']:
            sub_state_name = data['status_name']

        elif data['status_name'] == '已采购':
            if data['sub_status']==1:
                sub_state_name ='分开发货'
            else:
                sub_state_name= '已采购'

        else:
            sub_state_name = data['sub_status_name']

        sub_row = (sku['code'],
                   data['order_sn'],
                   sku['price'],
                   sku['quantity'],
                   float(sku['price']) * int(sku['quantity']),
                   sub_state_name,
                   sub_name,
                   sku['platform_url'],
                   sku['sku_info']['sku_img'],
                   # sku['sku_info']['specs'],
                   sku_spec,
                   sku['remark'],
                   data['status_name'],
                   data['status'],
                   sub_state_name,
                   data['sub_status']
                   )
        print(sub_row)
        sub_rows.append(sub_row)

    print(row)
    row_data = (row["order_sn"], row["status_name"], row["country_name"], row["created_time"],
                row["whare_house"], row["jiyun_packageNO"], row["platform_orderNo"], row["wuliu_order"],
                row["order_user"], row["purchaser"], row["order_platform"], row["address"], row["pay_status"],
                row["pay_order"], row['pay_time'], row["pay_method"], row["pay_amount"], row["delivery_fee"],
                row["total_payment"]
                )
    rows.append(row_data)

# print(rows)

# 创建一个新的工作簿
wb = Workbook()

# 获取活动工作表
ws = wb.active
# ws = wb.create_sheet("主订单数据表")

# 设置工作表标题
# ws.title = f"{country}-运费表"

# 写入数据到单元格 主表
# 写入数据到单元格
ws['A1'] = '主订单号'
ws['B1'] = '订单状态'
ws['C1'] = '寄送国家'
ws['D1'] = '创建时间'
ws['E1'] = '下单仓库'
ws['F1'] = '集运包裹号'
ws['G1'] = '平台订单号'
ws['H1'] = '采购物流单号'
ws['I1'] = '下单用户'
ws['J1'] = '采购员'
ws['K1'] = '下单平台'
ws['L1'] = '详细地址'
ws['M1'] = '支付状态'
ws['N1'] = '支付流水号'
ws['O1'] = '支付时间'
ws['P1'] = '支付方式'
ws['Q1'] = '支付金额'
ws['R1'] = '运费'
ws['S1'] = '订单总额'

# 写入多行数据
for row in rows:
    ws.append(row)

max_rows = ws.max_row  # 获取最大行
max_columns = ws.max_column
align = Alignment(horizontal='center', vertical='center')
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 35
ws.column_dimensions['G'].width = 15
ws.column_dimensions['J'].width = 15
ws.column_dimensions['K'].width = 15
ws.column_dimensions['L'].width = 15
ws.column_dimensions['M'].width = 24

for i in range(1, max_rows + 1):
    for j in range(1, max_columns + 1):
        if i > 1 and (j == 4 or j == 12 or j == 13):
            continue
        ws.cell(i, j).alignment = align

# 创建一个粗体字体样式
bold_font = Font(bold=True)
# 将粗体字体样式应用到 B1 单元格

for i in range(1, max_rows + 1):
    ws.cell(1, i).font = bold_font

# for j in range(1,max_columns+1):
#     ws.column_dimensions[j].width = 15


# 切换活动工作表，写入子订单数据
ws_sub = wb.create_sheet("Sheet2")
wb.active = ws_sub

# 写入数据到单元格
ws_sub['A1'] = '子订单id'
ws_sub['B1'] = '主订单号'
ws_sub['C1'] = '单价'
ws_sub['D1'] = '数量'
ws_sub['E1'] = '总货值'
ws_sub['F1'] = '子单状态'
ws_sub['G1'] = '商品名称'
ws_sub['H1'] = '商品链接'
ws_sub['I1'] = '主图链接'
ws_sub['J1'] = '商品规格'
ws_sub['K1'] = '子单备注'

ws_sub['L1'] = '主单状态'
ws_sub['M1'] = '主单状态码'
ws_sub['N1'] = '子单状态'
ws_sub['O1'] = '子单状态码'



# 写入多行数据到子单表
for sub_row in sub_rows:
    ws_sub.append(sub_row)

# 保存工作簿
# 将时间戳转换为本地时间元组
timestamp = time.time()
local_time = time.localtime(timestamp)

# # 获取日期时间
formatted_time = time.strftime("%Y-%m-%d_%H%M%S", local_time)
wb.save(f'haiou_daigou_orders_{formatted_time}.xlsx')

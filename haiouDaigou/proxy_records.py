import time
import requests

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

workbook = load_workbook(filename=r'0001.xlsx')
sheet_name = "代理数据"
ws = workbook[sheet_name]

max_rows = ws.max_row  # 获取最大行

proxy_ids = []

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
    for cell in row:
        # print(cell.value)
        proxy_ids.append(cell.value)

session = requests.session()

# 1 封装请求头
headers = {
    "Content-Type": "application/x-www-form-urlencoded;charset=UTF-8",
    "Origin": "https://jiyun-v4.haiouoms.com",
    "Referer": "https://jiyun-v4.haiouoms.com/",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36",
    "accept-language": "zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7,ja-JP;q=0.6,ja;q=0.5",
    "accept": "application/json, text/plain, */*",
    "language": "zh_CN",
    "authorization": "bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJodHRwczovL2FwaS5oYWlvdW9tcy5jb20vYXBpL2FkbWluL2xvZ2luIiwiaWF0IjoxNzQyMjA1MjAwLCJleHAiOjE3NDQ3OTcyMDAsIm5iZiI6MTc0MjIwNTIwMCwianRpIjoicWV4MVJZQkdZVE5ibDB6dSIsInN1YiI6IjQ0MDQiLCJwcnYiOiJkZjg4M2RiOTdiZDA1ZWY4ZmY4NTA4MmQ2ODZjNDVlODMyZTU5M2E5Iiwicm9sZSI6ImFkbWluIn0.1Jnld6HctTKjCGAlyToXDmpOaJWVmk35p_0dyuC7dcU"

    # "Accept-Encoding": "gzip, deflate, br, zstd"
}

session.headers = headers
invite_rows = []
agentIdMap = {}
deal_rows = []

# 获取代理数据
def getAllAgentData():
    page = 1,
    size = 1000
    data_agent = {
        "keyword": None,
        "page": page,
        "size": size
    }
    url = f"https://api-jiyun-v3.haiouoms.com/api/admin/agents?keyword=&page={page}&size={size}"
    response = session.get(data=data_agent, url=url)
    resp = response.json()
    for item in resp['data']:
        agentIdMap[item['user_id']]=item['id']
def getInvieRecords():
    data_invite = {
        "page": 1,
        "size": 10,
    }
    # 获取邀请记录
    for proxy_id in proxy_ids:
        url_invite = f"https://api-jiyun-v3.haiouoms.com/api/admin/users/{proxy_id}/invitations?page=1&size=10"
        response = session.get(url=url_invite, data=data_invite)
        res = response.json()

        if len(res['data']) > 0:
            for item in res['data']:
                if item['is_agent_invite'] == 1:
                    is_agent_invite = "是"
                else:
                    is_agent_invite = "否"
                row = {
                    "proxy_id": item['invite_id'],
                    "client_id": item['id'],
                    "client_nickName": item['name'],
                    "register_time": item['created_at'],
                    "lastLogin_time": item['last_login_at'],
                    "if_invited": is_agent_invite
                }
                row_data = (
                    row['proxy_id'], row['client_id'], row['client_nickName'], row['register_time'],
                    row['lastLogin_time'],
                    row['if_invited'])
                invite_rows.append(row_data)

    print(invite_rows)
def getDealRecords():
    # 处理成交记录
    # 获取成交记录
    data_deal = {
        'page': 1,
        'size': 10
    }

    for proxy_id in proxy_ids:
        agentId = agentIdMap[proxy_id]
        url_deal = f"https://api-jiyun-v3.haiouoms.com/api/admin/agents/{agentId}/deal-orders?page=1&size=10"
        # "https://api-jiyun-v3.haiouoms.com/api/admin/agents/27654/deal-orders?page=1&size=10"
        response = session.get(url=url_deal, data=data_deal)
        res = response.json()
        if len(res['data']) > 0:
            print(res['data'])
            for item in res['data']:
                row = {
                    "proxy_id": proxy_id,
                    "orderUser_id": item['user_id'],
                    "name": item['user_name'],
                    "order_number": item['order_number'],
                    "order_amount": item['order_amount'],
                    "commission_prop": item['proportion'],
                    "commission": item['commission_amount'],
                    "deal_time": item['created_at'],
                    "status": item['settled'],
                    "level": item['level']

                }
                row_data = (row['proxy_id'], row['orderUser_id'], row['name'],
                            row['order_number'], row['order_amount'], row['commission_prop'],
                            row['commission'],row['level'], row['deal_time'], row['status'])
                deal_rows.append(row_data)

getAllAgentData()

getInvieRecords()

getDealRecords()

# 写入邀请记录
ws_invite = workbook["邀请记录"]
for row in invite_rows:
    ws_invite.append(row)

# 写入成交记录
ws_deal = workbook["成交记录"]
for row in deal_rows:
    ws_deal.append(row)

# 保存工作簿
# 将时间戳转换为本地时间元组
timestamp = time.time()
local_time = time.localtime(timestamp)
formatted_time = time.strftime("%Y-%m-%d_%H%M%S", local_time)
workbook.save(f'proxy_data_{formatted_time}.xlsx')
